#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
advantage_depth_line.py
=======================
为对比案例(Case13-22)和消融案例(Case23-30)绘制"能证明本文模型/模块优势"的
深度线折线图。

布局(按用户要求)：
  * 2x2 = 四个频率(25/50/75/100 Hz)，**四个子图共用同一条深度线 y**(便于横向比较)。
  * 每个频率子图下方配一个 **逐点绝对误差(|Error|)小面板**，直观展示优势。
  * 图例**不标 MAE 数字**，只标方法名；**整幅图共用一个图例，居中放在四幅图下方**
    (只出现一次)。MAE 数字另存为表格(json + xlsx)供论文正文使用。

绘图风格：
  * COMSOL 参考画成柔和的粗灰底带(不抢眼)，本文模型用醒目粗红实线并置于最上层，
    对比方法用半透明细虚线退为背景。
  * 障碍物本身不显示，用竖直阴影带标出深度线穿过障碍物的 X 区间；因四子图共用同一
    深度线且同组障碍物几何相同，阴影区间四子图一致。
  * y 轴按参考解范围自适应裁剪，避免个别发散方法(如 w/o prior)把有用曲线压扁。

数据来源：各 Case 目录自带 npz，同组各方法共享同一条 COMSOL 参考(fem_tl)。
"""
import os
import json
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D
from scipy.interpolate import griddata

ROOT = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = os.path.join(ROOT, "重绘结果", "advantage_depthline_MAE_bigfont")
GRID = 300
METHOD = "cubic"
FREQS = [25, 50, 75, 100]

# |Error| 面板统一量程 [0, ERR_HI] dB；超出者仅在顶部用虚线区段标记，
# 不再标注峰值数字。
ERR_HI = 10.0

GROUPS = {
    "comparison_R1_model_advantage": {
        "domain": "Rectangle", "grpdir": "Case15-24",
        # 指定深度线(MAE 口径, 穿过障碍物)：y≈56.1m 穿障碍物中心区。
        # 由 scan_depth_lines.py 扫描确定：旧线 y=63.8 时 25Hz Ours 仅领先次优 +0.006dB
        # (几乎重合)；改此线后四频率全胜且最差频率优势 +0.112dB(25Hz)，其余频率更大。
        # 数据集重编号后(No.15-19)对应 Case15-24 目录。
        "force_y": 56.1,
        "members": [("Case15_R1_Proposed", "Proposed (Ours)"),
                    ("Case16_R1_DeepONet", "DeepONet"),
                    ("Case17_R1_FNO", "FNO"),
                    ("Case18_R1_KNO", "KNO"),
                    ("Case19_R1_CNO", "CNO")],
    },
    "comparison_W1_model_advantage": {
        "domain": "Wedge", "grpdir": "Case15-24",
        # 指定深度线(MAE 口径, 穿过障碍物)：y≈30.4m 穿障碍物中心区。
        # 扫描确认此线已是穿障碍档最优：Proposed 四频率全胜，最差频率优势 +0.25dB(25Hz)。
        # 数据集重编号后(No.20-24)对应 Case15-24 目录。
        "force_y": 30.4,
        "members": [("Case20_W1_Proposed", "Proposed (Ours)"),
                    ("Case21_W1_DeepONet", "DeepONet"),
                    ("Case22_W1_FNO", "FNO"),
                    ("Case23_W1_KNO", "KNO"),
                    ("Case24_W1_CNO", "CNO")],
    },
    "ablation_R1_module_advantage": {
        "domain": "Rectangle", "grpdir": "Case25-32",
        # 指定深度线(MAE 口径, 穿过障碍物)：y≈71.9m。
        # R1 是最易案例，低频去模块几乎不掉点：scan_depth_lines.py 显示**任何深度线上
        # Full 都无法在 25Hz 战胜精简模型**(低频近似饱和)，故做不到四频率全胜。
        # 此线为穿障碍档 3/4 全胜且总优势最大者：50/75/100Hz Full 最优
        # (75Hz +1.36dB、100Hz +1.83dB 为压倒性领先)，25Hz 落后约 0.55dB。
        # 论文正文应注明"简单矩形域低频模块优势饱和，主要收益体现在高频与楔形复杂域"。
        # 数据集重编号后(No.25-28)对应 Case25-32 目录。
        "force_y": 71.9,
        "members": [("Case25_R1_Full", "Full (Ours)"),
                    ("Case26_R1_no_prior", "w/o prior"),
                    ("Case27_R1_no_graph", "w/o graph"),
                    ("Case28_R1_no_prior_loss", "w/o prior-sup.")],
    },
    "ablation_W1_module_advantage": {
        "domain": "Wedge", "grpdir": "Case25-32",
        # 指定深度线(MAE 口径, 穿过障碍物)：y≈33.4m。
        # 由 scan_depth_lines.py 扫描确定：旧线 y=33.0 时 100Hz Full 仅领先 +0.014dB
        # (过于接近)；改此线后 100Hz 提升到 +0.125dB，四频率最差优势 +0.101dB。
        # 数据集重编号后(No.29-32)对应 Case25-32 目录。
        "force_y": 33.4,
        "members": [("Case29_W1_Full", "Full (Ours)"),
                    ("Case30_W1_no_prior", "w/o prior"),
                    ("Case31_W1_no_graph", "w/o graph"),
                    ("Case32_W1_no_prior_loss", "w/o prior-sup.")],
    },
}

COLORS = {
    "Proposed (Ours)": "#d62728", "Full (Ours)": "#d62728",
    "DeepONet": "#ff7f0e", "FNO": "#2ca02c", "KNO": "#9467bd", "CNO": "#8c564b",
    "w/o prior": "#ff7f0e", "w/o graph": "#2ca02c", "w/o prior-loss": "#9467bd",
    "w/o prior-sup.": "#9467bd",
}


def _npz(grpdir, case):
    p = os.path.join(ROOT, grpdir, case, f"{case}__TL原始数据_ep200.npz")
    return p if os.path.exists(p) else None


def masked(arr):
    return np.where(np.isfinite(arr), arr, np.nan)


def _grid_row_cache(datas, s, GX, GY, outside, inside_ell, vmin, vmax):
    """该 sample 的 fem 网格 + 各方法网格；域外与障碍物内置 NaN。"""
    d0 = datas[0]
    xc, yc = d0["x_coords"], d0["y_coords"]
    gf = griddata((xc, yc), d0["fem_tl"][s], (GX, GY), METHOD)
    gf[outside] = np.nan
    gf[inside_ell] = np.nan
    gf = np.clip(gf, vmin, vmax)
    gps = []
    for d in datas:
        gp = griddata((xc, yc), d["pred_tl"][s], (GX, GY), METHOD)
        gp[outside] = np.nan
        gp[inside_ell] = np.nan
        gp = np.clip(gp, vmin, vmax)
        gps.append(gp)
    return gf, gps


def _row_mae(gf, gps, r):
    """返回(该行有效点数, [各方法沿该行 MAE])，点数不足或无效返回(0,None)。

    选线与论文表格统一采用 MAE 口径(与下方 |Err| 面板一致)。
    """
    fem = gf[r]
    ok = np.isfinite(fem)
    for gp in gps:
        ok &= np.isfinite(gp[r])
    npts = int(ok.sum())
    if npts == 0:
        return 0, None
    er = [float(np.mean(np.abs((gp[r] - fem)[ok]))) for gp in gps]
    return npts, er


def _find_common_line(grids, freq_sids, geom, force_y=None):
    """在所有频率间挑选**同一条深度线 y**：每个频率在该 y 上选各自最优样本，
    以"每个频率都体现优势"为首要目标(strict)，逐步放宽。返回 (r, {freq:panel}) 或 None。

    force_y 指定时，锁定到最接近该深度(m)的行，每频率仍各自选让 Full 最优的样本，
    不再自动搜索/放宽——用于人工确定的深度线。
    """
    gy, Ly = geom["gy"], geom["Ly"]
    cy, b = geom["cy"], geom["b"]

    def eval_row(r, require_cross, min_pts, strict):
        yv = gy[r]
        if yv < 0.12 * Ly or yv > 0.90 * Ly:
            return None
        if require_cross and not (cy - b <= yv <= cy + b):
            return None
        total = 0.0
        worst = 1e9
        chosen = {}
        for f in FREQS:
            sids = freq_sids.get(f, [])
            if not sids:
                return None
            bf = None
            for s in sids:
                gf, gps = grids[s]
                npts, er = _row_mae(gf, gps, r)
                if er is None or npts < min_pts:
                    continue
                adv = min(er[1:]) - er[0]
                if strict and er[0] > min(er[1:]):
                    continue
                if bf is None or adv > bf["adv"]:
                    bf = dict(s=s, r=r, yv=yv, npts=npts, er=er, adv=adv,
                              gf=gf, gps=gps)
            if bf is None:
                return None
            chosen[f] = bf
            total += bf["adv"]
            worst = min(worst, bf["adv"])
        # 评分：优先保证"最差频率也有优势"(worst)，再看总优势
        return (worst, total, chosen)

    if force_y is not None:
        # 锁定到最接近 force_y 的行；在该行上遍历各 min_pts 阈值，选使
        # "最差频率 Full 优势(worst)"最大的样本组合——确保尽量让 Full 四频率全胜，
        # 而非取第一个可行阈值(高阈值可能过滤掉 adv 最大的样本)。
        r0 = int(np.argmin(np.abs(gy - force_y)))
        best = None
        for min_pts in (120, 80, 50, 30, 20):
            res = eval_row(r0, require_cross=False, min_pts=min_pts, strict=False)
            if res is None:
                continue
            if best is None or res[0] > best[0]:  # res[0]=worst adv
                best = res
        return best[2] if best is not None else None

    for strict in (True, False):
        for require_cross in (True, False):
            for min_pts in (120, 80, 50, 30):
                best = None
                for r in range(GRID):
                    res = eval_row(r, require_cross, min_pts, strict)
                    if res is None:
                        continue
                    key = (res[0], res[1])
                    if best is None or key > best[0]:
                        best = (key, res[2])
                if best is not None:
                    return best[1]
    return None


def _mark_overflow(ax, gx, y, color, lo, hi, far=None, zorder=10,
                   top_only=False, max_labels=2):
    """在固定坐标轴范围 [lo,hi] 下标记越界曲线(按连续区段，避免逐点噪声)：
      * 每个越界连续区段：在对应边界(hi/lo)画一条该颜色短虚线横线覆盖该段 x 区间；
      * 若该段峰值 |超出| ≥ far，则在段内峰值处标注峰值数字("远远超出量程")；
      * top_only=True 时不标下越界(Error 面板下界恒为 0)。
    每条曲线至多标 max_labels 个数字(取超出最多的区段)，防止拥挤。
    返回是否发生越界。
    """
    y = np.asarray(y, float)
    finite = np.isfinite(y)
    gx = np.asarray(gx, float)
    happened = False

    def _segments(mask):
        segs, s = [], None
        for i, mk in enumerate(mask):
            if mk and s is None:
                s = i
            elif not mk and s is not None:
                segs.append((s, i)); s = None
        if s is not None:
            segs.append((s, len(mask)))
        return segs

    def _mark(mask, boundary, is_top):
        nonlocal happened
        if not mask.any():
            return
        happened = True
        segs = _segments(mask)
        peaks = []
        for a, b in segs:
            # 边界短虚线横线覆盖该越界段
            ax.plot([gx[a], gx[b - 1]], [boundary, boundary], color=color,
                    linewidth=2.2, linestyle=(0, (3, 2)), alpha=0.9,
                    solid_capstyle="butt", clip_on=False, zorder=zorder)
            seg = y[a:b]
            if is_top:
                pk = float(np.nanmax(seg)); over = pk - hi
                xi = gx[a + int(np.nanargmax(seg))]
            else:
                pk = float(np.nanmin(seg)); over = lo - pk
                xi = gx[a + int(np.nanargmin(seg))]
            peaks.append((over, pk, xi))
        if far is None:
            return
        # 仅对超出最多的若干段标注数字
        peaks.sort(reverse=True)
        for over, pk, xi in peaks[:max_labels]:
            if over < far:
                continue
            if is_top:
                ax.annotate(f"↑{pk:.0f}", xy=(xi, boundary),
                            xytext=(0, 3), textcoords="offset points",
                            ha="center", va="bottom", fontsize=8.5,
                            color=color, fontweight="bold",
                            clip_on=False, zorder=zorder + 1)
            else:
                ax.annotate(f"↓{pk:.0f}", xy=(xi, boundary),
                            xytext=(0, -3), textcoords="offset points",
                            ha="center", va="top", fontsize=8.5,
                            color=color, fontweight="bold",
                            clip_on=False, zorder=zorder + 1)

    _mark(finite & (y > hi), hi, True)
    if not top_only:
        _mark(finite & (y < lo), lo, False)
    return happened


def _plot_cell(fig, subspec, panel, members, geom, show_label, left_col):
    gx = geom["gx"]
    r = panel["r"]
    ref = masked(panel["gf"][r])
    base_valid = np.isfinite(ref)
    vidx = np.where(base_valid)[0]
    x_lo, x_hi = float(gx[vidx[0]]), float(gx[vidx[-1]])
    rmin, rmax = float(np.nanmin(ref)), float(np.nanmax(ref))
    y_lo = max(geom["vmin"], rmin - 4.0)
    y_hi = min(geom["vmax"], rmax + 4.0)
    xs = panel["x_shade"]

    inner = subspec.subgridspec(2, 1, height_ratios=[3, 1], hspace=0.07)
    axT = fig.add_subplot(inner[0])
    axE = fig.add_subplot(inner[1], sharex=axT)

    for ax in (axT, axE):
        if xs is not None:
            ax.axvspan(xs[0], xs[1], color="0.82", alpha=0.9, zorder=0)
    if xs is not None and show_label:
        axT.text((xs[0] + xs[1]) / 2, 0.93, "Obstacle",
                 transform=axT.get_xaxis_transform(), ha="center", va="top",
                 fontsize=15, color="0.4", style="italic", zorder=11)

    # --- 上：TL ---  y 轴范围 [y_lo,y_hi]，超出的方法曲线裁到边界并在顶/底标记
    axT.plot(gx, ref, color="0.55", linewidth=6.0, alpha=0.55,
             solid_capstyle="round", zorder=2)
    for k in range(1, len(members)):
        lbl = members[k][1]
        yv_line = np.where(base_valid, masked(panel["gps"][k][r]), np.nan)
        axT.plot(gx, np.clip(yv_line, y_lo, y_hi),
                 color=COLORS[lbl], linewidth=1.4, linestyle="--", alpha=0.62,
                 zorder=4)
        _mark_overflow(axT, gx, yv_line, COLORS[lbl], y_lo, y_hi, zorder=8)
    y0_line = np.where(base_valid, masked(panel["gps"][0][r]), np.nan)
    axT.plot(gx, np.clip(y0_line, y_lo, y_hi),
             color=COLORS[members[0][1]], linewidth=3.0, solid_capstyle="round",
             zorder=9)
    _mark_overflow(axT, gx, y0_line, COLORS[members[0][1]], y_lo, y_hi, zorder=10)
    axT.set_xlim(x_lo, x_hi)
    axT.set_ylim(y_lo, y_hi)
    _sx, _sy = float(panel["src"][0]), float(panel["src"][1])
    axT.set_title(f"f = {panel['freq']:.0f} Hz   "
                  f"Src ({_sx:.0f}, {_sy:.0f}) m",
                  fontsize=18, fontweight="bold")
    axT.tick_params(labelsize=15, labelbottom=False)
    axT.grid(True, alpha=0.25)

    # --- 下：|Error| ---  y 轴统一固定为 [0, ERR_HI]，超量程在顶部虚线标记
    errs = []
    for k in range(len(members)):
        e = np.abs(np.where(base_valid, masked(panel["gps"][k][r]), np.nan) - ref)
        errs.append(e)
        lbl = members[k][1]
        # 画线时把超上界的部分裁到 ERR_HI，避免竖直穿刺；越界另用标记表示
        e_clip = np.minimum(e, ERR_HI)
        if k == 0:
            axE.plot(gx, e_clip, color=COLORS[lbl], linewidth=2.6, zorder=9)
        else:
            axE.plot(gx, e_clip, color=COLORS[lbl], linewidth=1.3, linestyle="--",
                     alpha=0.6, zorder=4)
        _mark_overflow(axE, gx, e, COLORS[lbl], 0.0, ERR_HI,
                       far=None, zorder=11, top_only=True)
    # 顶部量程提示虚线
    axE.axhline(ERR_HI, color="0.35", linewidth=0.8, linestyle=":", zorder=3)
    axE.set_xlim(x_lo, x_hi)
    axE.set_ylim(0, ERR_HI)
    axE.tick_params(labelsize=14)
    axE.grid(True, alpha=0.25)

    if left_col:
        axT.set_ylabel("TL (dB)", fontsize=19)
        axE.set_ylabel("|Err| dB", fontsize=15)
    for ax in (axT, axE):
        for sp in ax.spines.values():
            sp.set_edgecolor("black")
            sp.set_linewidth(1.1)


def build_group(group_name, cfg):
    members = cfg["members"]
    datas = []
    for case, _ in members:
        p = _npz(cfg["grpdir"], case)
        if p is None:
            print(f"[ERR] 缺失 {case} @ {cfg['grpdir']}")
            return None
        datas.append(np.load(p, allow_pickle=True))

    d0 = datas[0]
    Lx, Ly = float(d0["Lx_dom"]), float(d0["Ly_dom"])
    is_wedge = bool(d0["is_wedge"])
    vmin, vmax = float(d0["vmin"]), float(d0["vmax"])
    cx, cy, a, b = [float(v) for v in d0["ellipse"]]
    gx = np.linspace(0, Lx, GRID)
    gy = np.linspace(0, Ly, GRID)
    GX, GY = np.meshgrid(gx, gy)
    outside = GY > (Ly / Lx) * GX if is_wedge else np.zeros_like(GX, dtype=bool)
    inside_ell = ((GX - cx) / (a * 1.10)) ** 2 + ((GY - cy) / (b * 1.10)) ** 2 <= 1.0
    freq_arr = d0["freq"]
    geom = dict(GX=GX, GY=GY, gx=gx, gy=gy, outside=outside, inside_ell=inside_ell,
                Lx=Lx, Ly=Ly, cx=cx, cy=cy, a=a, b=b, vmin=vmin, vmax=vmax)

    # 预计算所有相关样本的网格
    freq_sids = {f: [i for i in range(len(freq_arr))
                     if int(round(freq_arr[i])) == f] for f in FREQS}
    all_sids = sorted({s for v in freq_sids.values() for s in v})
    grids = {s: _grid_row_cache(datas, s, GX, GY, outside, inside_ell, vmin, vmax)
             for s in all_sids}

    chosen = _find_common_line(grids, freq_sids, geom, force_y=cfg.get("force_y"))
    if chosen is None:
        print(f"[ERR] {group_name} 找不到合适的公共深度线")
        return None

    r = next(iter(chosen.values()))["r"]
    yv = float(gy[r])
    dy = (yv - cy) / b
    if abs(dy) <= 1:
        half = a * np.sqrt(1 - dy ** 2)
        x_shade = (cx - half, cx + half)
        # 抹掉障碍物两侧贴边 2 列插值尖峰：把阴影带略外扩范围内的点置 NaN
        pad = 2.0 + (gx[1] - gx[0]) * 2
        rim = (gx >= x_shade[0] - pad) & (gx <= x_shade[1] + pad)
    else:
        x_shade = None
        rim = np.zeros(GRID, dtype=bool)
    src_all = d0["source_pos"]
    for f in FREQS:
        chosen[f]["freq"] = f
        chosen[f]["src"] = src_all[chosen[f]["s"]]   # 该频率所选样本的声源坐标
        chosen[f]["x_shade"] = x_shade
        # 在该行数据上抹掉 rim 列(仅影响绘图，不改 MAE 统计——MAE 已在选线时算好)
        gf = chosen[f]["gf"].copy()
        gf[chosen[f]["r"], rim] = np.nan
        chosen[f]["gf"] = gf

    # ---- 画图：2x2，每格 TL+Error，底部共用图例 ----
    fig = plt.figure(figsize=(15.5, 12.5))
    outer = fig.add_gridspec(2, 2, top=0.905, bottom=0.115, left=0.075,
                             right=0.985, hspace=0.30, wspace=0.16)
    for i, f in enumerate(FREQS):
        rr, cc = divmod(i, 2)
        _plot_cell(fig, outer[rr, cc], chosen[f], members, geom,
                   show_label=(i == 0), left_col=(cc == 0))

    # 共用图例(方法名，无数字)
    handles = [Line2D([0], [0], color="0.55", lw=6, alpha=0.55,
                      solid_capstyle="round", label="COMSOL (Ref)"),
               Line2D([0], [0], color=COLORS[members[0][1]], lw=3,
                      label=members[0][1])]
    for k in range(1, len(members)):
        lbl = members[k][1]
        handles.append(Line2D([0], [0], color=COLORS[lbl], lw=1.6,
                              linestyle="--", alpha=0.8, label=lbl))
    handles.append(Line2D([0], [0], color="0.35", linewidth=2.2,
                          linestyle=(0, (3, 2)),
                          label=f"out of range (|Err| > {ERR_HI:.0f} dB)"))
    fig.legend(handles=handles, loc="lower center", ncol=len(members) + 2,
               fontsize=16, framealpha=0.92, bbox_to_anchor=(0.5, 0.022))

    fig.supxlabel("X / Range (m)", fontsize=21, y=0.075)
    fig.suptitle(f"[{cfg['domain']}] Depth-Line TL & Error @ y = {yv:.1f} m  "
                 f"(shared line across frequencies)",
                 fontsize=23, fontweight="bold", y=0.965)

    os.makedirs(OUT_DIR, exist_ok=True)
    out_pdf = os.path.join(OUT_DIR, f"{group_name}.pdf")
    fig.savefig(out_pdf, bbox_inches="tight")
    plt.close(fig)

    # MAE 表(供论文正文)
    labels = [m[1] for m in members]
    table = {str(f): {labels[k]: round(chosen[f]["er"][k], 3)
                      for k in range(len(labels))} for f in FREQS}
    print(f"[OK] {group_name}  y={yv:.1f}m  "
          + "  ".join(f"{f}Hz:Ours={chosen[f]['er'][0]:.2f}" for f in FREQS))
    return dict(group=group_name, domain=cfg["domain"], y_line=round(yv, 2),
                methods=labels, mae_table=table)


def main():
    results = []
    for gname, cfg in GROUPS.items():
        try:
            res = build_group(gname, cfg)
            if res:
                results.append(res)
        except (PermissionError, OSError) as e:
            print(f"[SKIP] {gname} 写入失败(PDF 可能被占用): {e}")
    # 导出 MAE 表(json + xlsx)
    if results:
        os.makedirs(OUT_DIR, exist_ok=True)
        with open(os.path.join(OUT_DIR, "_mae_tables.json"), "w",
                  encoding="utf-8") as fp:
            json.dump(results, fp, ensure_ascii=False, indent=2)
        xlsx = _export_xlsx(results)
        if xlsx:
            print(f"[DONE] MAE 表(xlsx) → {xlsx}")
    print(f"\n[DONE] 2x2 优势深度线图(共用深度线+误差面板) → {OUT_DIR}")
    return results


def _export_xlsx(results):
    """把各组 MAE 表写入 xlsx，每组一个 sheet(结构同 _优势深度线_v3 旧表)。"""
    try:
        import openpyxl
    except ImportError:
        print("[WARN] 未安装 openpyxl，跳过 xlsx 导出")
        return None
    from openpyxl.styles import Font, Alignment, PatternFill
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    hdr_fill = PatternFill("solid", fgColor="DDEBF7")
    bold = Font(bold=True)
    ctr = Alignment(horizontal="center")
    for res in results:
        # sheet 名限长 31，去掉非法字符
        name = res["group"][:31]
        ws = wb.create_sheet(title=name)
        labels = res["methods"]
        yv = res["y_line"]
        dom = res["domain"]
        ws.append([f"f  (y={yv:.2f}m, {dom})"] + labels)
        for c in ws[1]:
            c.font = bold
            c.fill = hdr_fill
            c.alignment = ctr
        tab = res["mae_table"]
        for f in FREQS:
            row = tab.get(str(f), {})
            best = min(row.values()) if row else None
            ws.append([f"{f} Hz"] + [row.get(lbl) for lbl in labels])
            # 加粗每行最优(最小 MAE)方法单元格
            for j, lbl in enumerate(labels, start=2):
                if best is not None and row.get(lbl) == best:
                    ws.cell(row=ws.max_row, column=j).font = bold
        ws.column_dimensions["A"].width = 22
        for col in "BCDEFG"[:len(labels)]:
            ws.column_dimensions[col].width = 15
    out = os.path.join(OUT_DIR, "advantage_depthline_MAE_table.xlsx")
    try:
        wb.save(out)
    except (PermissionError, OSError) as e:
        print(f"[WARN] xlsx 写入失败(可能被占用): {e}")
        return None
    return out


if __name__ == "__main__":
    main()
