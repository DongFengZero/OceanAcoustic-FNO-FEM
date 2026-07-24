# -*- coding: utf-8 -*-
"""Case43-50 推理时间性能分析表。
   Sheet1: Case43(Rect)/Case44(Wedge) COMSOL vs 1/2/4 GPU(A800) 加速对比 (4频率基准, 1 epoch)
   Sheet2: Case45-50 域尺寸缩放推理 (单频100Hz, 200 epoch训练日志)
   数据源: 补充实验3
"""
import os, re, glob
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = r'D:\Data'
SUP3 = os.path.join(BASE, '补充实验3')
OUT  = os.path.join(BASE, 'Case43-50')
os.makedirs(OUT, exist_ok=True)

def read(fp):
    with open(fp, encoding='utf-8', errors='ignore') as f: return f.read()

def get_test_block(txt):
    """从'推理时间统计摘要'里取 测试集 平均每样本ms、每epoch s、总测试s"""
    m = re.search(r'测试集:\s*\n\s*平均每样本时间:\s*([\d.]+)\s*ms\s*\n\s*平均每epoch时间:\s*([\d.]+)\s*s\s*\n\s*总测试时间:\s*([\d.]+)\s*s', txt)
    if m: return float(m.group(1)), float(m.group(2)), float(m.group(3))
    return None, None, None

def get_throughput_test(txt):
    """测试集总体统计里的吞吐量(第2个'吞吐量')"""
    tps = re.findall(r'吞吐量:\s*([\d.]+)\s*samples/s', txt)
    return float(tps[1]) if len(tps) >= 2 else (float(tps[0]) if tps else None)

# ---- Case43/44: COMSOL + 1/2/4 GPU ----
# COMSOL 全局均值 ms/sample 和吞吐
def comsol_mean(txt):
    m = re.search(r'全局每样本求解时间统计.*?均值\s*:\s*([\d.]+)\s*ms', txt, re.S)
    tp = re.search(r'平均吞吐\s*:\s*([\d.]+)\s*samples/s', txt)
    return (float(m.group(1)) if m else None), (float(tp.group(1)) if tp else None)

bench = []  # (case,dataset,geom,N,method,ms_per_sample,throughput,epoch_s)
CFG = [
    (43, 'R1', 'Rect.', 21737, 'Rect', 'run_rectangle_*.log'),
    (44, 'W1', 'Wedge', 10680, 'Wedge', 'run_wedge_*.log'),
]
for no, did, geom, N, sub, runglob in CFG:
    # COMSOL
    ctxt = read(glob.glob(os.path.join(SUP3, sub, runglob))[0])
    cms, ctp = comsol_mean(ctxt)
    bench.append([no, did, geom, N, 'COMSOL (CPU)', cms, ctp, None])
    # GPU 1/2/4
    for g in (1, 2, 4):
        gf = glob.glob(os.path.join(SUP3, sub, f'gpu{g}_full_run_*.log'))[0]
        gtxt = read(gf)
        ms, ep, tot = get_test_block(gtxt)
        tp = get_throughput_test(gtxt)
        bench.append([no, did, geom, N, f'{g}×A800 GPU', ms, tp, ep])

# ---- Case45-50: 域尺寸缩放 ----
SCALE = [
    (45, 'R4', 'Rect.', 128, 128, 'Case45_R4'),
    (46, 'R5', 'Rect.', 256, 256, 'Case46_R5'),
    (47, 'R6', 'Rect.', 512, 512, 'Case47_R6'),
    (48, 'W4', 'Wedge', 128, 128, 'Case48_W4'),
    (49, 'W5', 'Wedge', 256, 256, 'Case49_W5'),
    (50, 'W6', 'Wedge', 512, 512, 'Case50_W6'),
]
scale_rows = []
for no, did, geom, lx, ly, folder in SCALE:
    lf = glob.glob(os.path.join(SUP3, folder, '*.log'))[0]
    txt = read(lf)
    ms, ep, tot = get_test_block(txt)
    tp = get_throughput_test(txt)
    # 节点数(从 periodic_tl_raw 保存行 或 网格行)
    mnode = re.search(r'节点[ :]*(\d{3,})', txt)
    N = int(mnode.group(1)) if mnode else None
    scale_rows.append([no, did, geom, lx, ly, N, ms, tp, ep])

# ---- 写 xlsx ----
thin=Side(style='thin',color='BBBBBB'); bd=Border(left=thin,right=thin,top=thin,bottom=thin)
ctr=Alignment(horizontal='center',vertical='center',wrap_text=True)
hf=PatternFill('solid',fgColor='305496'); hfont=Font(bold=True,color='FFFFFF',size=10)
cmsl=PatternFill('solid',fgColor='FCE4D6'); rfill=PatternFill('solid',fgColor='DDEBF7'); wfill=PatternFill('solid',fgColor='E2EFDA')

wb=Workbook()
# Sheet1
ws=wb.active; ws.title='COMSOL_vs_GPU加速'
ws.merge_cells('A1:I1')
ws['A1']='Case 43-44 推理时间性能对比 · COMSOL(CPU) vs 神经算子(A800 GPU, 1/2/4卡) — R1/W1几何, 25/50/75/100Hz, 测试集800样本'
ws['A1'].font=Font(bold=True,size=12)
ws.merge_cells('A2:I2')
ws['A2']='说明: 单样本时间与吞吐量为测试集推理性能。COMSOL为CPU有限元求解(全频率合并均值)，GPU为A800上神经算子前向推理。加速比=GPU吞吐/COMSOL吞吐。多GPU为数据并行，单样本时延近似不变而吞吐/墙钟随卡数近线性提升。'
ws['A2'].font=Font(italic=True,size=9,color='595959'); ws.row_dimensions[2].height=30
hdr=['Case','数据集','几何','网格节点数','方法/硬件','单样本时间(ms)','吞吐量(samp/s)','每epoch墙钟(s)','加速比 vs COMSOL']
ws.append([None]*9)  # row3 spacer
for j,h in enumerate(hdr,1):
    c=ws.cell(4,j,h); c.fill=hf; c.font=hfont; c.alignment=ctr; c.border=bd
# comsol throughput per case for speedup
cms_tp={}
for row in bench:
    if row[4].startswith('COMSOL'): cms_tp[row[0]]=row[6]
rr=5
for row in bench:
    no,did,geom,N,method,ms,tp,ep=row
    sp = (tp/cms_tp[no]) if (tp and cms_tp.get(no)) else None
    vals=[no,did,geom,N,method,ms,tp,ep, (round(sp,1) if sp else '—')]
    fill = cmsl if method.startswith('COMSOL') else (wfill if geom=='Wedge' else rfill)
    for j,v in enumerate(vals,1):
        c=ws.cell(rr,j, v if v is not None else '—'); c.alignment=ctr; c.border=bd; c.fill=fill
        if j in (6,7,8) and isinstance(v,float): c.number_format='0.00'
    rr+=1
for i,w in enumerate([6,8,8,12,14,14,14,15,16],1):
    ws.column_dimensions[get_column_letter(i)].width=w
ws.freeze_panes='A5'

# Sheet2
ws2=wb.create_sheet('域尺寸缩放推理')
ws2.merge_cells('A1:H1')
ws2['A1']='Case 45-50 推理时间 · 域尺寸缩放 (R4/R5/R6/W4/W5/W6, 100Hz, 单GPU测试集)'
ws2['A1'].font=Font(bold=True,size=12)
ws2.merge_cells('A2:H2')
ws2['A2']='说明: 取自各案例200轮训练日志的测试集推理时间摘要。随域尺寸(128→256→512)增大，节点数与单样本推理时间显著上升。'
ws2['A2'].font=Font(italic=True,size=9,color='595959'); ws2.row_dimensions[2].height=26
hdr2=['Case','数据集','几何','Lx','Ly','网格节点数','单样本时间(ms)','吞吐量(samp/s)']
ws2.append([None]*8)
for j,h in enumerate(hdr2,1):
    c=ws2.cell(4,j,h); c.fill=hf; c.font=hfont; c.alignment=ctr; c.border=bd
rr=5
for row in scale_rows:
    no,did,geom,lx,ly,N,ms,tp,ep=row
    fill=wfill if geom=='Wedge' else rfill
    vals=[no,did,geom,lx,ly,N,ms,tp]
    for j,v in enumerate(vals,1):
        c=ws2.cell(rr,j, v if v is not None else '—'); c.alignment=ctr; c.border=bd; c.fill=fill
        if j in (7,8) and isinstance(v,float): c.number_format='0.00'
    rr+=1
for i,w in enumerate([6,8,8,6,6,12,14,14],1):
    ws2.column_dimensions[get_column_letter(i)].width=w
ws2.freeze_panes='A5'

out=os.path.join(OUT,'Case43-50_推理时间性能分析.xlsx')
wb.save(out)
print('saved:',out)
for r in bench: print('bench',r)
print('---')
for r in scale_rows: print('scale',r)
