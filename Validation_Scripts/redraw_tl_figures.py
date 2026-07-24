#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
redraw_tl_figures.py
====================
重绘 Case1-12 / 13-22 / 23-30 / 31-36 的 TL 样本图，解决原图字体过小的问题。

基于每个 case 目录内自带的 `*__TL原始数据_ep200.npz`（由 restore_tl_figure.py
所用的同一数据），重新绘制：
  1) 样本对比图  <case>__TL对比_重绘.pdf   —— 每个样本单独一页(1x3: Ours|COMSOL|Error)，字体放大
  2) 深度线对比图 <case>__深度线对比_重绘.pdf —— 取中间深度 y=Ly/2 的横切线，1x2:
        左: 沿线 TL-vs-Range 曲线 (Ours vs COMSOL 叠加)
        右: Ours 2D 场图并标出该深度线位置
  3) 误差数据    深度线误差汇总.xlsx        —— 不画在图上，用表格呈现

原始 case 目录保持不变（作为备份），所有输出写入独立的新文件夹。
"""
import os
import glob
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.patches import Ellipse as MplEllipse
from matplotlib.backends.backend_pdf import PdfPages
from mpl_toolkits.axes_grid1 import make_axes_locatable
from scipy.interpolate import griddata
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

ROOT = os.path.dirname(os.path.abspath(__file__))
OUT_ROOT = os.path.join(ROOT, "重绘结果")
GRID_RES = 250
METHOD = "cubic"


def build_grid(data, sample_idx, grid_res=GRID_RES, method=METHOD):
    """复刻 restore_tl_figure 的插值+遮罩逻辑，返回 (GX, GY, grid_pred, grid_fem, meta)."""
    xc = data["x_coords"]
    yc = data["y_coords"]
    Lx = float(data["Lx_dom"])
    Ly = float(data["Ly_dom"])
    is_wedge = bool(data["is_wedge"])
    ellipse = data["ellipse"]
    vmin = float(data["vmin"])
    vmax = float(data["vmax"])

    gx = np.linspace(0, Lx, grid_res)
    gy = np.linspace(0, Ly, grid_res)
    GX, GY = np.meshgrid(gx, gy)

    if is_wedge:
        outside = GY > (Ly / Lx) * GX
    else:
        outside = np.zeros_like(GX, dtype=bool)

    gp = griddata((xc, yc), data["pred_tl"][sample_idx], (GX, GY), method=method)
    gf = griddata((xc, yc), data["fem_tl"][sample_idx], (GX, GY), method=method)

    has_ellipse = ellipse.size == 4
    if has_ellipse:
        cx, cy, a, b = [float(v) for v in ellipse]
        inside = ((GX - cx) / a) ** 2 + ((GY - cy) / b) ** 2 <= 1.0
        gp[inside] = np.nan
        gf[inside] = np.nan
    gp[outside] = np.nan
    gf[outside] = np.nan

    gp = np.clip(gp, vmin, vmax)
    gf = np.clip(gf, vmin, vmax)

    meta = dict(Lx=Lx, Ly=Ly, is_wedge=is_wedge, ellipse=ellipse,
                vmin=vmin, vmax=vmax,
                domain_label=str(data["domain_label"]),
                epoch=int(data["epoch"]),
                source_pos=data["source_pos"][sample_idx],
                freq=float(data["freq"][sample_idx]))
    return gx, gy, GX, GY, gp, gf, meta


def _decorate(ax, meta, mark_line_y=None):
    Lx, Ly = meta["Lx"], meta["Ly"]
    if meta["is_wedge"]:
        ax.plot([0, Lx], [0, Ly], "k-", linewidth=1.8)
        ax.plot([Lx, Lx], [0, Ly], color="gray", linewidth=1.2, linestyle="--")
    if meta["ellipse"].size == 4:
        cx, cy, a, b = [float(v) for v in meta["ellipse"]]
        ax.add_patch(MplEllipse((cx, cy), width=2 * a, height=2 * b,
                                fill=False, edgecolor="k", linewidth=1.8))
    sp = meta["source_pos"]
    ax.plot(sp[0], sp[1], "r*", markersize=16)
    if mark_line_y is not None:
        ax.axhline(mark_line_y, color="magenta", linewidth=2.2, linestyle="--")


def draw_sample_page(pdf, gx, gy, GX, GY, gp, gf, meta, sample_idx, n_total):
    """每个样本一页：1x3 Ours | COMSOL | Error，字体放大适合展示。"""
    Lx, Ly = meta["Lx"], meta["Ly"]
    vmin, vmax = meta["vmin"], meta["vmax"]
    extent = (0, Lx, Ly, 0)
    err = np.abs(gp - gf)
    avg_err = float(np.nanmean(err))
    err_vmax = min(float(np.nanmax(err)) if np.any(np.isfinite(err)) else 10.0, 10.0)
    sp = meta["source_pos"]
    freq = meta["freq"]

    fig, axes = plt.subplots(1, 3, figsize=(20, 7))
    titles = [
        f"Ours TL  (f={freq:.0f} Hz)\nSrc: ({sp[0]:.1f}, {sp[1]:.1f})",
        f"COMSOL TL  (f={freq:.0f} Hz)",
        f"Error vs COMSOL | Avg: {avg_err:.2f} dB",
    ]
    dl = [gp, gf, err]
    cml = ["jet", "jet", "Reds"]
    vminl = [vmin, vmin, 0]
    vmaxl = [vmax, vmax, err_vmax]
    lbl = ["TL (dB)", "TL (dB)", "Error (dB)"]

    for j in range(3):
        ax = axes[j]
        im = ax.imshow(dl[j], extent=extent, origin="upper", cmap=cml[j],
                       aspect="equal", vmin=vminl[j], vmax=vmaxl[j])
        _decorate(ax, meta)
        ax.set_xlabel("X / Range (m)", fontsize=16)
        ax.set_ylabel("Y / Depth (m)", fontsize=16)
        ax.set_title(titles[j], fontsize=17, fontweight="bold")
        ax.tick_params(labelsize=13)
        ax.set_xlim(0, Lx)
        ax.set_ylim(Ly, 0)
        cbar = plt.colorbar(im, ax=ax, fraction=0.046, pad=0.04, shrink=0.9)
        cbar.set_label(lbl[j], fontsize=14)
        cbar.ax.tick_params(labelsize=12)
        for spine in ax.spines.values():
            spine.set_edgecolor("black")
            spine.set_linewidth(1.5)

    fig.suptitle(f"[{meta['domain_label']}] TL Field  Epoch {meta['epoch']}  "
                 f"| Sample {sample_idx + 1}/{n_total}",
                 fontsize=19, fontweight="bold")
    plt.tight_layout(rect=[0, 0, 1, 0.93])
    pdf.savefig(fig, bbox_inches="tight")
    plt.close(fig)


def choose_depth_row(gx, gy, gp, gf, meta):
    """选取深度线所在行：避开椭圆障碍物所在的 y 带，并选覆盖有效点最多的一行。

    返回 (row_index, y_value)。
    """
    Lx, Ly = meta["Lx"], meta["Ly"]
    # 椭圆遮挡的 y 带（含安全裕量）
    forbidden = None
    if meta["ellipse"].size == 4:
        cx, cy, a, b = [float(v) for v in meta["ellipse"]]
        margin = 0.06 * Ly
        forbidden = (cy - b - margin, cy + b + margin)

    valid_counts = (np.isfinite(gp) & np.isfinite(gf)).sum(axis=1)  # 每行有效点数
    max_count = valid_counts.max()
    if max_count == 0:
        return int(len(gy) // 2), float(gy[len(gy) // 2])

    # 偏好深度：在障碍物上方、约 0.35*Ly 处（浅一些，声学展示常用），
    # 若落在禁带内则移到禁带上沿之上。
    target = 0.35 * Ly
    if forbidden is not None and forbidden[0] <= target <= forbidden[1]:
        target = max(forbidden[0] - 0.02 * Ly, 0.12 * Ly)

    floor = 0.12 * Ly   # 深度下限：避开海面(y≈0)退化行
    # 覆盖充分性阈值：至少覆盖最大值的 40%，且不少于 30 个点
    min_cov = max(0.4 * max_count, 30)

    # 在“避开障碍物 + 非贴海面 + 覆盖充分”的行里，选最靠近目标深度的一行。
    best_row, best_dist = None, None
    for r, yv in enumerate(gy):
        if yv < floor:
            continue
        if forbidden is not None and forbidden[0] <= yv <= forbidden[1]:
            continue  # 跳过穿过障碍物的行
        if valid_counts[r] < min_cov:
            continue  # 覆盖太少（如楔形深处几乎全在域外）
        dist = abs(yv - target)
        if best_dist is None or dist < best_dist:
            best_dist, best_row = dist, r

    if best_row is None:  # 兜底：取覆盖最多且不贴海面的行
        cand = [r for r, yv in enumerate(gy) if yv >= floor]
        best_row = max(cand, key=lambda r: valid_counts[r]) if cand \
            else int(np.argmax(valid_counts))
    return best_row, float(gy[best_row])


def draw_depth_line(out_pdf, gx, gy, GX, GY, gp, gf, meta):
    """深度线对比图：避开障碍物的横切线，1x2。返回沿线数据用于误差表。

    - 深度线避开椭圆障碍物所在 y 带。
    - 曲线只展示有效数据段（以右侧有数据处为起点，域外/障碍物内不展示）。
    - 左右两图等高，色标高度与场图一致；字体加大。
    """
    Lx, Ly = meta["Lx"], meta["Ly"]
    vmin, vmax = meta["vmin"], meta["vmax"]

    row, y_actual = choose_depth_row(gx, gy, gp, gf, meta)

    line_pred = gp[row]     # [grid_res] 沿 x
    line_fem = gf[row]
    x_axis = gx

    valid = np.isfinite(line_pred) & np.isfinite(line_fem)
    diff = line_pred - line_fem
    abs_err = np.abs(diff)

    # 只展示有效数据段：从最左有效点到最右有效点（去掉域外/障碍物空段）
    if valid.any():
        vidx = np.where(valid)[0]
        x_lo, x_hi = float(x_axis[vidx[0]]), float(x_axis[vidx[-1]])
    else:
        x_lo, x_hi = 0.0, Lx

    fig, axes = plt.subplots(1, 2, figsize=(19, 8),
                             gridspec_kw={"width_ratios": [1, 1]})

    # 左：沿线 TL-vs-Range 曲线 (Ours vs COMSOL)，仅有效段
    ax = axes[0]
    ax.plot(x_axis[valid], line_fem[valid], color="#1f77b4", linewidth=3.0,
            label="COMSOL (Ref)")
    ax.plot(x_axis[valid], line_pred[valid], color="#d62728", linewidth=2.6,
            linestyle="--", label="Ours")
    ax.set_xlabel("X / Range (m)", fontsize=19)
    ax.set_ylabel("TL (dB)", fontsize=19)
    ax.set_title(f"Depth-Line TL @ y={y_actual:.1f} m  (f={meta['freq']:.0f} Hz)",
                 fontsize=20, fontweight="bold")
    ax.tick_params(labelsize=16)
    ax.grid(True, alpha=0.3)
    ax.legend(fontsize=17, loc="best")
    ax.set_xlim(x_lo, x_hi)
    ax.set_box_aspect(1.0)  # 与右侧等宽高方框，保证两图等高
    for spine in ax.spines.values():
        spine.set_edgecolor("black")
        spine.set_linewidth(1.5)

    # 右：Ours 2D 场图并标出深度线（色标高度=图高）
    ax = axes[1]
    extent = (0, Lx, Ly, 0)
    im = ax.imshow(gp, extent=extent, origin="upper", cmap="jet",
                   aspect="equal", vmin=vmin, vmax=vmax)
    _decorate(ax, meta, mark_line_y=y_actual)
    ax.set_xlabel("X / Range (m)", fontsize=19)
    ax.set_ylabel("Y / Depth (m)", fontsize=19)
    ax.set_title(f"Ours TL Field  (depth line @ y={y_actual:.1f} m)",
                 fontsize=20, fontweight="bold")
    ax.tick_params(labelsize=16)
    ax.set_xlim(0, Lx)
    ax.set_ylim(Ly, 0)
    # 色标高度与场图严格一致
    divider = make_axes_locatable(ax)
    cax = divider.append_axes("right", size="5%", pad=0.12)
    cbar = plt.colorbar(im, cax=cax)
    cbar.set_label("TL (dB)", fontsize=17)
    cbar.ax.tick_params(labelsize=14)
    for spine in ax.spines.values():
        spine.set_edgecolor("black")
        spine.set_linewidth(1.5)

    fig.suptitle(f"[{meta['domain_label']}] Depth-Line Comparison  Epoch {meta['epoch']}",
                 fontsize=22, fontweight="bold")
    plt.tight_layout(rect=[0, 0, 1, 0.95])
    fig.savefig(out_pdf, bbox_inches="tight")
    plt.close(fig)

    # 误差统计（仅有效点）
    if valid.sum() > 0:
        d = diff[valid]
        ae = abs_err[valid]
        stats = dict(
            y_line=y_actual,
            freq=meta["freq"],
            src_x=float(meta["source_pos"][0]),
            src_y=float(meta["source_pos"][1]),
            n_valid=int(valid.sum()),
            rmse=float(np.sqrt(np.mean(d ** 2))),
            mae=float(np.mean(ae)),
            max_abs=float(np.max(ae)),
            mean_bias=float(np.mean(d)),
            std_err=float(np.std(d)),
        )
    else:
        stats = dict(y_line=y_actual, freq=meta["freq"],
                     src_x=float(meta["source_pos"][0]),
                     src_y=float(meta["source_pos"][1]),
                     n_valid=0, rmse=np.nan, mae=np.nan, max_abs=np.nan,
                     mean_bias=np.nan, std_err=np.nan)

    line_data = dict(x=x_axis, pred=line_pred, fem=line_fem,
                     diff=diff, abs_err=abs_err, valid=valid)
    return stats, line_data


# ---- 分组：对比案例 / 消融案例，用于绘制"所有案例对比"的深度线汇总图 ----
GROUPS = [
    {"name": "对比_R1_模型对比",
     "cases": [("Case13_R1_Proposed", "Proposed (Ours)"),
               ("Case14_R1_DeepONet", "DeepONet"),
               ("Case15_R1_FNO", "FNO"),
               ("Case16_R1_KNO", "KNO"),
               ("Case17_R1_CNO", "CNO")]},
    {"name": "对比_W1_模型对比",
     "cases": [("Case18_W1_Proposed", "Proposed (Ours)"),
               ("Case19_W1_DeepONet", "DeepONet"),
               ("Case20_W1_FNO", "FNO"),
               ("Case21_W1_KNO", "KNO"),
               ("Case22_W1_CNO", "CNO")]},
    {"name": "消融_R1_模块消融",
     "cases": [("Case23_R1_Full", "Full (Ours)"),
               ("Case24_R1_no_prior", "no_prior"),
               ("Case25_R1_no_graph", "no_graph"),
               ("Case26_R1_no_prior_loss", "no_prior_loss")]},
    {"name": "消融_W1_模块消融",
     "cases": [("Case27_W1_Full", "Full (Ours)"),
               ("Case28_W1_no_prior", "no_prior"),
               ("Case29_W1_no_graph", "no_graph"),
               ("Case30_W1_no_prior_loss", "no_prior_loss")]},
]


def _find_case_npz(case):
    hits = glob.glob(os.path.join(ROOT, "Case*", case, "*__TL原始数据_ep200.npz"))
    return hits[0] if hits else None


def draw_group_depth_line(out_pdf, group, grid_res=GRID_RES, method=METHOD):
    """把一组案例(对比/消融)在同一条深度线上的 Ours 曲线叠加到一张图，
    与共享的 COMSOL 参考对比。返回逐案例误差统计 list。

    组内几何/障碍物/代表样本一致 → 共用同一条深度线与同一 COMSOL 参考。
    """
    members = group["cases"]
    # 用第一个成员(通常是 Proposed/Full)确定深度线与网格
    ref_npz = _find_case_npz(members[0][0])
    if ref_npz is None:
        print(f"[SKIP-GROUP] {group['name']}: 找不到参考案例 {members[0][0]}")
        return []
    ref_data = np.load(ref_npz, allow_pickle=True)
    idx0 = representative_index(ref_data)
    gx, gy, GX, GY, gp0, gf0, meta0 = build_grid(ref_data, idx0, grid_res, method)
    row, y_line = choose_depth_row(gx, gy, gp0, gf0, meta0)
    fem_line = gf0[row]                      # 共享 COMSOL 参考
    base_valid = np.isfinite(fem_line)

    Lx, Ly = meta0["Lx"], meta0["Ly"]
    cmap = plt.get_cmap("tab10")

    fig, ax = plt.subplots(1, 1, figsize=(13, 9))
    stats_rows = []
    all_valid = base_valid.copy()
    ax.plot(gx[base_valid], fem_line[base_valid], color="black", linewidth=3.6,
            label="COMSOL (Ref)", zorder=10)

    for k, (case, label) in enumerate(members):
        npz = _find_case_npz(case)
        if npz is None:
            print(f"[SKIP-GROUP-MEMBER] {case} 缺失")
            continue
        d = np.load(npz, allow_pickle=True)
        idx = representative_index(d)
        _, _, _, _, gp, gf, _ = build_grid(d, idx, grid_res, method)
        pred_line = gp[row]
        v = np.isfinite(pred_line) & base_valid
        all_valid &= np.isfinite(pred_line)
        ax.plot(gx[v], pred_line[v], color=cmap(k % 10), linewidth=2.4,
                linestyle="--", label=label)
        # 误差统计（沿线，仅共同有效点）
        vv = np.isfinite(pred_line) & np.isfinite(fem_line)
        if vv.any():
            diff = (pred_line - fem_line)[vv]
            ae = np.abs(diff)
            stats_rows.append(dict(
                group=group["name"], case=case, label=label,
                y_line=y_line, freq=meta0["freq"], n_valid=int(vv.sum()),
                rmse=float(np.sqrt(np.mean(diff ** 2))),
                mae=float(np.mean(ae)), max_abs=float(np.max(ae)),
                mean_bias=float(np.mean(diff))))

    # x 轴仅展示有效数据段
    if base_valid.any():
        vidx = np.where(base_valid)[0]
        ax.set_xlim(float(gx[vidx[0]]), float(gx[vidx[-1]]))
    ax.set_xlabel("X / Range (m)", fontsize=19)
    ax.set_ylabel("TL (dB)", fontsize=19)
    ax.set_title(f"[{meta0['domain_label']}] Depth-Line Comparison @ y={y_line:.1f} m  "
                 f"(f={meta0['freq']:.0f} Hz)", fontsize=20, fontweight="bold")
    ax.tick_params(labelsize=16)
    ax.grid(True, alpha=0.3)
    ax.legend(fontsize=16, loc="best", framealpha=0.9)
    for spine in ax.spines.values():
        spine.set_edgecolor("black")
        spine.set_linewidth(1.5)
    plt.tight_layout()
    fig.savefig(out_pdf, bbox_inches="tight")
    plt.close(fig)
    return stats_rows


def representative_index(data):
    """选代表样本：优先最高频率的第一个样本。"""
    freqs = data["freq"]
    target = np.max(freqs)
    idx = int(np.where(freqs == target)[0][0])
    return idx


def process_case(npz_path):
    case = os.path.basename(os.path.dirname(npz_path))
    data = np.load(npz_path, allow_pickle=True)
    n = data["pred_tl"].shape[0]

    out_dir = os.path.join(OUT_ROOT, case)
    os.makedirs(out_dir, exist_ok=True)

    # 1) 样本对比图（多页 PDF，每样本一页）
    sample_pdf = os.path.join(out_dir, f"{case}__TL对比_重绘.pdf")
    with PdfPages(sample_pdf) as pdf:
        for i in range(n):
            gx, gy, GX, GY, gp, gf, meta = build_grid(data, i)
            draw_sample_page(pdf, gx, gy, GX, GY, gp, gf, meta, i, n)

    # 2) 深度线对比图（代表样本一张）
    idx = representative_index(data)
    gx, gy, GX, GY, gp, gf, meta = build_grid(data, idx)
    line_pdf = os.path.join(out_dir, f"{case}__深度线对比_重绘.pdf")
    stats, line_data = draw_depth_line(line_pdf, gx, gy, GX, GY, gp, gf, meta)
    stats["case"] = case
    stats["domain"] = meta["domain_label"]
    stats["rep_sample"] = idx + 1
    return case, stats, line_data


def write_excel(all_stats, all_lines, group_stats=None):
    xlsx = os.path.join(OUT_ROOT, "深度线误差汇总.xlsx")
    wb = openpyxl.Workbook()

    # 汇总表
    ws = wb.active
    ws.title = "误差汇总"
    headers = ["Case", "Domain", "代表样本#", "频率(Hz)", "声源X(m)", "声源Y(m)",
               "深度线Y(m)", "有效点数", "RMSE(dB)", "MAE(dB)",
               "最大绝对误差(dB)", "平均偏差(dB)", "误差标准差(dB)"]
    ws.append(headers)
    hdr_fill = PatternFill("solid", fgColor="4472C4")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = hdr_fill
    for s in all_stats:
        ws.append([s["case"], s["domain"], s["rep_sample"], round(s["freq"], 0),
                   round(s["src_x"], 2), round(s["src_y"], 2),
                   round(s["y_line"], 2), s["n_valid"],
                   round(s["rmse"], 4), round(s["mae"], 4),
                   round(s["max_abs"], 4), round(s["mean_bias"], 4),
                   round(s["std_err"], 4)])
    for col in ws.columns:
        w = max(len(str(c.value)) for c in col if c.value is not None) + 2
        ws.column_dimensions[col[0].column_letter].width = max(w, 10)
    ws.freeze_panes = "A2"

    # 分组对比(对比/消融)误差表：同一深度线上各案例 vs COMSOL
    if group_stats:
        wsg = wb.create_sheet(title="分组对比误差", index=1)
        ghdr = ["分组", "Case", "方法/标签", "深度线Y(m)", "频率(Hz)", "有效点数",
                "RMSE(dB)", "MAE(dB)", "最大绝对误差(dB)", "平均偏差(dB)"]
        wsg.append(ghdr)
        for c in wsg[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.fill = hdr_fill
        for s in group_stats:
            wsg.append([s["group"], s["case"], s["label"], round(s["y_line"], 2),
                        round(s["freq"], 0), s["n_valid"], round(s["rmse"], 4),
                        round(s["mae"], 4), round(s["max_abs"], 4),
                        round(s["mean_bias"], 4)])
        for col in wsg.columns:
            w = max(len(str(c.value)) for c in col if c.value is not None) + 2
            wsg.column_dimensions[col[0].column_letter].width = max(w, 10)
        wsg.freeze_panes = "A2"

    # 每个 case 的逐点沿线数据
    for case, ld in all_lines.items():
        sheet = case[:31]  # excel sheet 名 <=31 字符
        wsd = wb.create_sheet(title=sheet)
        wsd.append(["X_Range(m)", "Ours_TL(dB)", "COMSOL_TL(dB)",
                    "Diff(Ours-COMSOL)(dB)", "AbsError(dB)", "Valid"])
        for c in wsd[1]:
            c.font = Font(bold=True)
        x, p, f, df, ae, v = (ld["x"], ld["pred"], ld["fem"],
                              ld["diff"], ld["abs_err"], ld["valid"])
        for k in range(len(x)):
            wsd.append([
                round(float(x[k]), 3),
                None if not np.isfinite(p[k]) else round(float(p[k]), 4),
                None if not np.isfinite(f[k]) else round(float(f[k]), 4),
                None if not np.isfinite(df[k]) else round(float(df[k]), 4),
                None if not np.isfinite(ae[k]) else round(float(ae[k]), 4),
                bool(v[k]),
            ])
    wb.save(xlsx)
    return xlsx


def main():
    os.makedirs(OUT_ROOT, exist_ok=True)
    npz_files = sorted(glob.glob(os.path.join(ROOT, "Case*", "Case*",
                                              "*__TL原始数据_ep200.npz")))
    print(f"发现 {len(npz_files)} 个 case npz 文件")
    all_stats, all_lines, failed = [], {}, []
    for npz in npz_files:
        try:
            case, stats, line_data = process_case(npz)
        except (PermissionError, OSError) as e:
            cname = os.path.basename(os.path.dirname(npz))
            failed.append((cname, str(e)))
            print(f"[SKIP] {cname}  写入失败(文件可能被占用): {e}")
            continue
        all_stats.append(stats)
        all_lines[case] = line_data
        print(f"[OK] {case}  RMSE={stats['rmse']:.3f}dB MAE={stats['mae']:.3f}dB "
              f"@y={stats['y_line']:.1f}m")
    if failed:
        print(f"\n[WARN] {len(failed)} 个 case 未完成(请关闭 PDF 查看器后重跑): "
              f"{[c for c, _ in failed]}")

    # 分组"所有案例对比"深度线图（对比案例 + 消融案例）
    group_dir = os.path.join(OUT_ROOT, "_分组对比深度线")
    os.makedirs(group_dir, exist_ok=True)
    group_stats = []
    for g in GROUPS:
        out_pdf = os.path.join(group_dir, f"{g['name']}__深度线汇总.pdf")
        try:
            rows = draw_group_depth_line(out_pdf, g)
            group_stats.extend(rows)
            print(f"[OK-GROUP] {g['name']}  ({len(rows)} 个案例叠加)")
        except (PermissionError, OSError) as e:
            print(f"[SKIP-GROUP] {g['name']} 写入失败(文件可能被占用): {e}")

    xlsx = write_excel(all_stats, all_lines, group_stats)
    print(f"\n[DONE] 图输出 → {OUT_ROOT}")
    print(f"[DONE] 误差表 → {xlsx}")


if __name__ == "__main__":
    main()
